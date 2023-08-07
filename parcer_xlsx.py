import openpyxl


def get_data_xlsx():
    wb = openpyxl.load_workbook("Средний балл 2023.xlsx", read_only=True, data_only=True)

    sheets = wb.sheetnames[:4]

    data = []
    name_in_data = []

    for i in sheets:
        current_sheet = wb[i]

        for row in range(2, current_sheet.max_row + 1):
            if current_sheet[row][1].value is None:
                break
            # print(current_sheet.title, current_sheet[row][1].value, current_sheet[row][27].value, current_sheet[row][28].value)
            FIO_split = current_sheet[row][1].value.split()
            FIO = f"{FIO_split[0]} {FIO_split[1]} {FIO_split[2]}"
            average_mark = current_sheet[row][27].value
            birthday = current_sheet[row][28].value

            if FIO not in name_in_data:
                abbiturient = {
                    'FIO': FIO,
                    'Average mark': "",
                    'Birthday': "",
                    'INFO': 0,
                    'EKONOM': 0,
                    'POVAR': 0,
                    'EKOLOG': 0,
                }

                match current_sheet.title:
                    case "Информац системы и программиров":
                        abbiturient["INFO"] = 1
                        if average_mark != "#DIV/0!" and average_mark is not None:
                            abbiturient['Average mark'] = round(average_mark, 3)

                        if birthday is not None:
                            abbiturient["Birthday"] = birthday
                        print(sheets[1:])
                        for j in sheets[1:]:
                            sheet = wb[j]

                            for r in range(2, sheet.max_row + 1):
                                if sheet[r][1].value is None:
                                    break

                                fio_split = sheet[r][1].value.split()
                                fio = f"{fio_split[0]} {fio_split[1]} {fio_split[2]}"
                                mark = sheet[r][27].value
                                brthd = sheet[r][28].value
                                if fio == abbiturient['FIO']:
                                    match sheet.title:
                                        case "Экономика и БУ":
                                            abbiturient['EKONOM'] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd

                                        case "Поварское и кондитерское дело":
                                            abbiturient['POVAR'] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd

                                        case "экологическая безопасность":
                                            abbiturient['EKOLOG'] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd

                                        case _:
                                            print()
                                    break

                    case "Экономика и БУ":
                        abbiturient["EKONOM"] = 1
                        if average_mark != "#DIV/0!" and average_mark is not None:
                            abbiturient['Average mark'] = round(average_mark, 3)

                        if birthday is not None:
                            abbiturient["Birthday"] = birthday
                        print(sheets[:1] + sheets[2:])
                        for j in sheets[:1] + sheets[2:]:
                            sheet = wb[j]
                            for r in range(2, sheet.max_row + 1):
                                if sheet[r][1].value is None:
                                    break

                                fio_split = sheet[r][1].value.split()
                                fio = f"{fio_split[0]} {fio_split[1]} {fio_split[2]}"
                                mark = sheet[r][27].value
                                brthd = sheet[r][28].value

                                if fio == abbiturient['FIO']:
                                    match sheet.title:
                                        case "Информац системы и программиров":
                                            abbiturient["INFO"] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd

                                        case "Поварское и кондитерское дело":
                                            abbiturient["POVAR"] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd

                                        case "экологическая безопасность":
                                            abbiturient["EKOLOG"] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd

                                        case _:
                                            print()
                                    break

                    case "Поварское и кондитерское дело":
                        abbiturient["POVAR"] = 1
                        if average_mark != "#DIV/0!" and average_mark is not None:
                            abbiturient['Average mark'] = round(average_mark, 3)

                        if birthday is not None:
                            abbiturient["Birthday"] = birthday
                        print(sheets[:2] + sheets[3:])
                        for j in sheets[:2] + sheets[3:]:
                            sheet = wb[j]
                            for r in range(2, sheet.max_row + 1):
                                if sheet[r][1].value is None:
                                    break

                                fio_split = sheet[r][1].value.split()
                                fio = f"{fio_split[0]} {fio_split[1]} {fio_split[2]}"
                                mark = sheet[r][27].value
                                brthd = sheet[r][28].value

                                if fio == abbiturient["FIO"]:
                                    match sheet.title:
                                        case "Информац системы и программиров":
                                            abbiturient["INFO"] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd

                                        case "Экономика и БУ":
                                            abbiturient["EKONOM"] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd

                                        case "экологическая безопасность":
                                            abbiturient["EKOLOG"] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd
                                        case _:
                                            print()

                    case "экологическая безопасность":
                        abbiturient["EKOLOG"] = 1
                        if average_mark != "#DIV/0!" and average_mark is not None:
                            abbiturient['Average mark'] = round(average_mark, 3)

                        if birthday is not None:
                            abbiturient["Birthday"] = birthday
                        print(sheets[:3])
                        for j in sheets[:3]:
                            sheet = wb[j]
                            for r in range(2, sheet.max_row + 1):
                                if sheet[r][1].value is None:
                                    break

                                fio_split = sheet[r][1].value.split()
                                fio = f"{fio_split[0]} {fio_split[1]} {fio_split[2]}"
                                mark = sheet[r][27].value
                                brthd = sheet[r][28].value

                                if fio == abbiturient["FIO"]:
                                    match sheet.title:
                                        case "Информац системы и программиров":
                                            abbiturient["INFO"] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd

                                        case "Экономика и БУ":
                                            abbiturient["EKONOM"] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd

                                        case "Поварское и кондитерское дело":
                                            abbiturient["POVAR"] = 1
                                            if mark != "#DIV/0!" and mark is not None:
                                                abbiturient['Average mark'] = round(mark, 3)
                                            if brthd is not None:
                                                abbiturient['Birthday'] = brthd
                                        case _:
                                            print()

                    case _:
                        print()
                if abbiturient["Birthday"] != "" and abbiturient['Average mark'] != "":
                    name_in_data.append(FIO)
                    data.append(abbiturient)



            # if FIO in name_in_data:
            #     for name in data:
            #         if name["FIO"] == FIO:
            #             match current_sheet.title:
            #                 case "Информац системы и программиров":
            #                     name["INFO"] = 1
            #                     if average_mark != "#DIV/0!" and average_mark is not None:
            #                         name['Average mark'] = round(average_mark, 3)
            #
            #                 case "Экономика и БУ":
            #                     name["EKONOM"] = 1
            #                     if average_mark != "#DIV/0!" and average_mark is not None:
            #                         name['Average mark'] = round(average_mark, 3)
            #
            #                 case "Поварское и кондитерское дело":
            #                     name["POVAR"] = 1
            #                     if average_mark != "#DIV/0!" and average_mark is not None:
            #                         name['Average mark'] = round(average_mark, 3)
            #
            #                 case "экологическая безопасность":
            #                     name["EKOLOG"] = 1
            #                     if average_mark != "#DIV/0!" and average_mark is not None:
            #                         name['Average mark'] = round(average_mark, 3)
            #
            #                 case _:
            #                     print()
            #             break
            # else:
            #     if average_mark is not None and birthday is not None:
            #         abbiturient = {
            #             'FIO': FIO,
            #             'Average mark': average_mark,
            #             'Birthday': birthday,
            #             'INFO': 0,
            #             'EKONOM': 0,
            #             'POVAR': 0,
            #             'EKOLOG': 0,
            #         }
            #
            #         match current_sheet.title:
            #             case "Информац системы и программиров":
            #                 abbiturient["INFO"] = 1
            #
            #             case "Экономика и БУ":
            #                 abbiturient["EKONOM"] = 1
            #
            #             case "Поварское и кондитерское дело":
            #                 abbiturient["POVAR"] = 1
            #
            #             case "экологическая безопасность":
            #                 abbiturient["EKOLOG"] = 1
            #
            #             case _:
            #                 print()
            #         name_in_data.append(FIO)
            #         data.append(abbiturient)

    print(*data, sep="\n")
    return data


def main():
    get_data_xlsx()


if __name__ == "__main__":
    main()